#!/usr/bin/env python3
import time

import openpyxl
from googletrans import Translator
import argparse

from tqdm import tqdm

# This const need for fix up  process of translate
# To be more specific:  after 250 line google don't give stable results
FIX_DELAY = 0.3


# Read the file and create xls
def create_translation_xlsx(input_filename, output_filename):

    # Initialization translator
    translator = Translator()

    # Create new XLSX file
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Read source file
    with open(input_filename, 'r', encoding='utf-8') as file:
        lines = file.readlines()
        doc_len = len(lines)-1
        progress_bar = tqdm(total=doc_len, desc="Traslate lines...", unit="line", unit_scale=1)

    # Iter for every string and add translate
    for i in range(0, doc_len):
        original_text = lines[i]
        translated_text = translator.translate(original_text, src='en', dest='ru').text

        sheet.cell(row=i+1, column=1, value=original_text)
        sheet.cell(row=i+1, column=2, value=translated_text)
        time.sleep(FIX_DELAY)
        progress_bar.update(1)

    # Save  XLSX file
    workbook.save(output_filename)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Translate text from an input file and save it to an XLSX file.")
    parser.add_argument("input_file", help="Input text file to translate")
    parser.add_argument("output_file", nargs='?', default="output.xlsx", help="Output XLSX file to save translations")
    args = parser.parse_args()

    input_file = args.input_file
    output_file = args.output_file

    create_translation_xlsx(input_file, output_file)
